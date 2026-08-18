from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import config
from utils.text_normalizer import normalizar_texto, normalizar_para_comparacao


class ErroPlanilha(Exception):
    pass


@dataclass
class MapaColunas:
    linha_cabecalho: int
    colunas: dict = field(default_factory=dict)

    def letra(self, alvo: str):
        return self.colunas.get(alvo)


@dataclass
class DadosPlanilha:
    caminho: Path
    wb_template: object
    wb_valores: object
    aba_itens: object
    aba_baixas: object
    aba_resumo: object = None
    mapa_itens: MapaColunas = None
    mapa_baixas: MapaColunas = None
    df_itens: pd.DataFrame = None
    df_baixas: pd.DataFrame = None
    lancamentos: list = field(default_factory=list)


def localizar_aba(wb, nomes_alvo: list):
    abas = {ws.title: ws for ws in wb.worksheets}
    for alvo in nomes_alvo:
        if alvo in abas:
            return abas[alvo]
    alvo_norm = normalizar_para_comparacao(nomes_alvo[0])
    for titulo, ws in abas.items():
        if normalizar_para_comparacao(titulo) == alvo_norm:
            return ws
    for titulo, ws in abas.items():
        if alvo_norm in normalizar_para_comparacao(titulo):
            return ws
    return None


def detectar_linha_cabecalho(ws, colunas_alvo: dict, max_linhas: int = None) -> int:
    max_linhas = max_linhas or config.MAX_LINHAS_VARREDURA_CABECALHO
    melhores = []
    for r in range(1, min(max_linhas, ws.max_row) + 1):
        achados = 0
        textos = [normalizar_texto(ws.cell(row=r, column=c).value)
                  for c in range(1, ws.max_column + 1)]
        for valores in colunas_alvo.values():
            for v in valores:
                vn = normalizar_para_comparacao(v)
                if any(vn and (vn == normalizar_para_comparacao(t) or vn in normalizar_para_comparacao(t)) for t in textos if t):
                    achados += 1
                    break
        melhores.append((achados, r))
    melhor = max(melhores, key=lambda x: x[0])
    if melhor[0] >= 2:
        return melhor[1]
    return 1


def mapear_colunas(ws, linha_cabecalho: int, colunas_alvo: dict) -> MapaColunas:
    mapa = MapaColunas(linha_cabecalho=linha_cabecalho)
    celulas = {
        ws.cell(row=linha_cabecalho, column=c).column_letter: normalizar_para_comparacao(
            ws.cell(row=linha_cabecalho, column=c).value
        )
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=linha_cabecalho, column=c).value is not None
    }
    for alvo, aliases in colunas_alvo.items():
        for alias in aliases:
            an = normalizar_para_comparacao(alias)
            if not an:
                continue
            for letra, t in celulas.items():
                if an == t:
                    mapa.colunas[alvo] = letra
                    break
            if alvo in mapa.colunas:
                break
    for alvo, aliases in colunas_alvo.items():
        if alvo in mapa.colunas:
            continue
        for alias in aliases:
            an = normalizar_para_comparacao(alias)
            if not an:
                continue
            for letra, t in celulas.items():
                if an in t:
                    mapa.colunas[alvo] = letra
                    break
            if alvo in mapa.colunas:
                break
    return mapa


def parsear_valor(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    t = str(valor).strip().replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return round(float(t), config.PRECISAO_VALOR)
    except ValueError:
        return 0.0


def _limpar_celula(v):
    if isinstance(v, str) and v.startswith("'"):
        return v[1:]
    return v


def _extrair_linhas(ws, mapa: MapaColunas, colunas_alvo: dict) -> list:
    linhas = []
    for r in range(mapa.linha_cabecalho + 1, ws.max_row + 1):
        valores = {}
        vazia = True
        for alvo in colunas_alvo:
            letra = mapa.letra(alvo)
            if letra:
                v = _limpar_celula(ws[f"{letra}{r}"].value)
                valores[alvo] = v
                if v is not None and str(v).strip() != "" and not str(v).lstrip().startswith("="):
                    vazia = False
        if not vazia:
            valores["linha"] = r
            linhas.append(valores)
    return linhas


def _garantir_colunas(df: pd.DataFrame, colunas: list) -> pd.DataFrame:
    for col in colunas:
        if col not in df.columns:
            df[col] = None
    return df[colunas]


def _str_codigo(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _preencher_derivados(df_itens: pd.DataFrame, df_baixas: pd.DataFrame) -> pd.DataFrame:
    if df_itens.empty:
        return df_itens
    df_itens["saldo_disponivel"] = pd.to_numeric(df_itens["saldo_disponivel"], errors="coerce")
    faltando = df_itens["saldo_disponivel"].isna()
    if not faltando.any():
        return df_itens
    qtd_baixada = pd.Series(0.0, index=df_itens.index)
    if df_baixas is not None and not df_baixas.empty:
        bq = pd.to_numeric(df_baixas.get("quantidade"), errors="coerce").fillna(0.0)
        bc = df_baixas.get("codigo").map(_str_codigo)
        soma = pd.DataFrame({"c": bc, "q": bq}).groupby("c")["q"].sum()
        qtd_baixada = df_itens["codigo_barras"].map(_str_codigo).map(soma).fillna(0.0)
    base = pd.to_numeric(df_itens.get("saldo_base"), errors="coerce")
    saldo_calc = (base - qtd_baixada).round(2)
    df_itens.loc[faltando, "saldo_disponivel"] = saldo_calc[faltando]
    return df_itens


def extrair_itens(ws, mapa: MapaColunas) -> pd.DataFrame:
    linhas = _extrair_linhas(ws, mapa, config.COLUNAS_ITENS)
    colunas = ["linha", "codigo_barras", "descricao", "preco_unit", "saldo_disponivel", "qtd_contratada", "saldo_base", "grupo", "num_item"]
    df = pd.DataFrame(linhas)
    if not df.empty:
        df = df.rename(columns={"saldo": "saldo_disponivel"})
    df = _garantir_colunas(df, colunas)
    df["preco_unit"] = df["preco_unit"].apply(parsear_valor)
    df["saldo_disponivel"] = pd.to_numeric(df["saldo_disponivel"], errors="coerce")
    return df.reset_index(drop=True)


def extrair_baixas(ws, mapa: MapaColunas) -> pd.DataFrame:
    linhas = _extrair_linhas(ws, mapa, config.COLUNAS_BAIXAS)
    colunas = ["linha", "data", "of", "codigo", "descricao", "quantidade", "preco_unit", "valor_total"]
    df = pd.DataFrame(linhas)
    df = _garantir_colunas(df, colunas)
    df["preco_unit"] = df["preco_unit"].apply(parsear_valor)
    df["valor_total"] = df["valor_total"].apply(parsear_valor)
    return df.reset_index(drop=True)


def carregar_planilha(origem) -> DadosPlanilha:
    if isinstance(origem, (str, Path)):
        caminho = Path(origem)
        if not caminho.exists():
            raise ErroPlanilha(f"Arquivo não encontrado: {caminho}")
        nome = caminho.name
        wb_template = load_workbook(caminho, data_only=False)
        wb_valores = load_workbook(caminho, data_only=True)
    else:
        origem.seek(0)
        wb_template = load_workbook(origem, data_only=False)
        origem.seek(0)
        wb_valores = load_workbook(origem, data_only=True)
        nome = getattr(origem, "name", "planilha.xlsx")
        caminho = Path(nome)

    aba_itens = localizar_aba(wb_template, [config.ABAS["itens"]])
    aba_baixas = localizar_aba(wb_template, [config.ABAS["baixas"]])
    aba_resumo = localizar_aba(wb_template, [config.ABAS["resumo"]])

    if aba_itens is None:
        raise ErroPlanilha(
            f"Aba '{config.ABAS['itens']}' não encontrada. Abas disponíveis: "
            f"{[ws.title for ws in wb_template.worksheets]}"
        )
    if aba_baixas is None:
        aba_baixas = wb_template.active

    aba_itens_valores = localizar_aba(wb_valores, [config.ABAS["itens"]])
    aba_baixas_valores = localizar_aba(wb_valores, [config.ABAS["baixas"]])

    linha_cab = detectar_linha_cabecalho(aba_itens, config.COLUNAS_ITENS)
    mapa_itens = mapear_colunas(aba_itens, linha_cab, config.COLUNAS_ITENS)

    linha_cab_b = detectar_linha_cabecalho(aba_baixas, config.COLUNAS_BAIXAS)
    mapa_baixas = mapear_colunas(aba_baixas, linha_cab_b, config.COLUNAS_BAIXAS)

    obrigatorias = ["descricao", "saldo"]
    ausentes = [c for c in obrigatorias if mapa_itens.letra(c) is None]
    if ausentes:
        raise ErroPlanilha(
            f"Colunas obrigatórias não encontradas na aba '{aba_itens.title}': {ausentes}. "
            f"Colunas detectadas: {mapa_itens.colunas}"
        )

    df_itens = extrair_itens(aba_itens_valores, mapa_itens)
    df_baixas = extrair_baixas(aba_baixas_valores, mapa_baixas)
    df_itens = _preencher_derivados(df_itens, df_baixas)

    return DadosPlanilha(
        caminho=caminho,
        wb_template=wb_template,
        wb_valores=wb_valores,
        aba_itens=aba_itens,
        aba_baixas=aba_baixas,
        aba_resumo=aba_resumo,
        mapa_itens=mapa_itens,
        mapa_baixas=mapa_baixas,
        df_itens=df_itens,
        df_baixas=df_baixas,
    )


def totais_contrato(dados: DadosPlanilha) -> dict:
    df = dados.df_itens
    if df.empty:
        return {"valor_total": 0.0, "saldo_total": 0.0}
    preco = pd.to_numeric(df["preco_unit"], errors="coerce").fillna(0.0)
    saldo = pd.to_numeric(df["saldo_disponivel"], errors="coerce").fillna(0.0)
    valor_total = float((preco * saldo).sum())
    saldo_total = float(saldo.sum())
    return {"valor_total": round(valor_total, 2), "saldo_total": round(saldo_total, 2)}


def salvar_template(dados: DadosPlanilha, destino) -> None:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    dados.wb_template.save(destino)


def exportar_bytes(dados: DadosPlanilha) -> bytes:
    from io import BytesIO

    buf = BytesIO()
    dados.wb_template.save(buf)
    return buf.getvalue()