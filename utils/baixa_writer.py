from dataclasses import dataclass
from datetime import date, datetime
import math

from openpyxl.utils import get_column_letter, range_boundaries


@dataclass
class Lancamento:
    data: str
    of: str
    codigo: str
    descricao: str
    quantidade: float
    preco_unit: float
    valor_total: float
    linha: int = 0
    grupo: str = ""
    num_item: str = ""
    gravado: bool = False


def aplicar_baixas(dados, escolhas, numero_of="OF"):
    lancamentos = []
    for esc in escolhas:
        linha = esc["linha"]
        qtd = float(esc["quantidade"])
        if qtd <= 0:
            continue
        idx = dados.df_itens.index[dados.df_itens["linha"] == linha]
        if len(idx) != 1:
            continue
        i = idx[0]
        saldo_atual = dados.df_itens.at[i, "saldo_disponivel"]
        if saldo_atual is not None and not (isinstance(saldo_atual, float) and math.isnan(saldo_atual)):
            saldo_num = float(saldo_atual)
            if saldo_num <= 0 or qtd > saldo_num:
                continue
            dados.df_itens.at[i, "saldo_disponivel"] = saldo_num - qtd
        preco = float(dados.df_itens.at[i, "preco_unit"] or 0.0)
        lancamentos.append(Lancamento(
            data=date.today().strftime("%d/%m/%Y"),
            of=numero_of,
            codigo=str(dados.df_itens.at[i, "codigo_barras"] or ""),
            descricao=str(dados.df_itens.at[i, "descricao"] or ""),
            quantidade=qtd,
            preco_unit=preco,
            valor_total=round(qtd * preco, 2),
            linha=linha,
            grupo=str(dados.df_itens.at[i, "grupo"] or ""),
            num_item=str(dados.df_itens.at[i, "num_item"] or ""),
        ))
    dados.lancamentos.extend(lancamentos)
    return lancamentos


def desfazer_baixas(dados, lancamentos):
    for lanc in lancamentos:
        lanc.gravado = False
        idx = dados.df_itens.index[dados.df_itens["linha"] == lanc.linha]
        if len(idx) == 1:
            i = idx[0]
            s = dados.df_itens.at[i, "saldo_disponivel"]
            if s is not None and not (isinstance(s, float) and math.isnan(s)):
                dados.df_itens.at[i, "saldo_disponivel"] = float(s) + lanc.quantidade
    dados.lancamentos = [l for l in dados.lancamentos if l not in lancamentos]


def _valor_celula(cell) -> float:
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _proxima_linha_livre(ws, linha_cab: int) -> int:
    for r in range(linha_cab + 1, ws.max_row + 2):
        vazia = True
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and (not isinstance(v, str) or not v.lstrip().startswith("=")):
                vazia = False
                break
        if vazia:
            return r
    return ws.max_row + 1


def _celula_estilo(ws, linha_cab: int, r: int, letra: str):
    for origem_r in range(r - 1, linha_cab, -1):
        if ws[f"{letra}{origem_r}"].value is not None:
            return ws[f"{letra}{origem_r}"]
    return None


def _formato_data(fmt: str) -> bool:
    f = (fmt or "").upper()
    return any(k in f for k in ("DD", "MM", "YYYY", "AAAA"))


def _estender_tabelas(ws, primeira_linha: int, n: int) -> None:
    if n <= 0:
        return
    ultima = primeira_linha + n - 1
    for tabela in list(ws.tables.values()):
        min_col, min_row, max_col, max_row = range_boundaries(tabela.ref)
        if max_row >= primeira_linha - 1 and min_row <= primeira_linha:
            nova_ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max(ultima, max_row)}"
            )
            tabela.ref = nova_ref


def gravar_no_template(dados, lancamentos) -> int:
    pendentes = [l for l in lancamentos if not l.gravado]
    if not pendentes:
        return 0

    saldo_letra = dados.mapa_itens.letra("saldo")
    if saldo_letra:
        for lanc in pendentes:
            cell = dados.aba_itens[f"{saldo_letra}{lanc.linha}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            atual = _valor_celula(cell) or 0.0
            cell.value = round(max(atual - lanc.quantidade, 0.0), 2)

    aba_baixas = dados.aba_baixas
    mapa = dados.mapa_baixas
    free = _proxima_linha_livre(aba_baixas, mapa.linha_cabecalho)
    for i, lanc in enumerate(pendentes):
        r = free + i
        for alvo, valor in (
            ("data", lanc.data),
            ("of", lanc.of),
            ("codigo", lanc.codigo),
            ("descricao", lanc.descricao),
            ("quantidade", lanc.quantidade),
            ("preco_unit", lanc.preco_unit),
            ("valor_total", lanc.valor_total),
            ("grupo", lanc.grupo),
            ("num_item", lanc.num_item),
        ):
            letra = mapa.letra(alvo)
            if not letra:
                continue
            origem = _celula_estilo(aba_baixas, mapa.linha_cabecalho, r, letra)
            cell = aba_baixas[f"{letra}{r}"]
            if isinstance(cell.value, str) and cell.value.lstrip().startswith("="):
                continue
            if alvo == "data" and origem is not None and _formato_data(origem.number_format):
                valor = datetime.strptime(lanc.data, "%d/%m/%Y").date()
            cell.value = valor
            if origem is not None:
                cell._style = origem._style
        lanc.gravado = True

    _estender_tabelas(aba_baixas, free, len(pendentes))
    return len(pendentes)